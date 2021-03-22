import openpyxl

i = 0
j = 0
list = []

wb =openpyxl.load_workbook("Python/excel/work.xlsx")
sheet = wb.worksheets[0]

def printCorrectOrder(i, j):
    gyou = i
    retsu = j
    while True:
        if sheet.cell(row=gyou, column=retsu).value == None:
            break
        print(sheet.cell(row=gyou, column=retsu).value)
        list.append(sheet.cell(row=gyou, column=retsu).value)
        gyou += 1

def writeUpsideOrder(i, j):
    gyou = i
    retsu = j
    cnt = len(list) - 1
    
    while cnt >= 0:
        print(list[cnt])
        sheet.cell(row=gyou, column=retsu).value= list[cnt]
        cnt -= 1
        gyou += 1
        #wb.save("Python/excel/work.xlsx")

def printUpsideOder(i, j):
    gyou = i
    retsu = j
    while True:
        if sheet.cell(row=gyou, column=retsu).value == None:
            break
        print(sheet.cell(row=gyou, column=retsu).value)
        gyou += 1

def main():
    i = 1
    j = 1
    printCorrectOrder(i, j)
    i = 1
    j = 2
    writeUpsideOrder(i, j)
    printUpsideOder(i, j)

main()